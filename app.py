from flask import Flask, render_template, request, redirect, send_file
from openpyxl import Workbook
import os
from datetime import datetime

app = Flask(__name__)
excel_file = "بيانات_الطلاب.xlsx"

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['student_name']
    guardian = request.form['guardian_name']
    birth_date = request.form['birth_date']
    card_number = request.form['card_number']
    phone = request.form['guardian_phone']
    bank = request.form['bank_name']

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["الاسم الثلاثي", "اسم ولي الأمر", "تاريخ الميلاد", "رقم بطاقة السكن", "هاتف ولي الأمر", "اسم المصرف"])
        wb.save(excel_file)

    from openpyxl import load_workbook
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([name, guardian, birth_date, card_number, phone, bank])
    wb.save(excel_file)

    return redirect('/success')

@app.route('/success')
def success():
    return render_template('success.html')

@app.route('/download')
def download():
    return send_file(excel_file, as_attachment=True)

if __name__ == '__main__':
    app.run()
